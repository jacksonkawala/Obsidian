#!/usr/bin/env python3
"""Extract TSF Lift 1 shift summaries from WhatsApp chat into Excel with color coding."""

import re
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

chat_file = sys.argv[1] if len(sys.argv) > 1 else "../raw/tsf/_chat.txt"

with open(chat_file, "r", encoding="utf-8") as f:
    lines = f.readlines()

# Reassemble messages: a new message starts with [DD/MM/YYYY, ...] or ‎[DD/MM/YYYY, ...]
msg_start = re.compile(r'^[\u200e]?\[(\d{2}/\d{2}/\d{4}),\s*(\d{1,2}:\d{2}:\d{2}\s*[AP]M)\]\s*([^:]+):\s*(.*)')
messages = []
current = None

for line in lines:
    m = msg_start.match(line)
    if m:
        if current:
            messages.append(current)
        current = {
            'date': m.group(1),
            'time': m.group(2),
            'sender': m.group(3).strip(),
            'body': m.group(4)
        }
    elif current:
        current['body'] += '\n' + line.rstrip('\n')

if current:
    messages.append(current)

# Find shift summary messages
summaries = []
for msg in messages:
    body = msg['body']
    if 'shift summary' not in body.lower() and 'Shift Summary' not in body:
        continue
    if 'TSF' not in body:
        continue
    summaries.append(msg)

print(f"Found {len(summaries)} TSF Shift Summary messages")

# Parse each summary
rows = []
for msg in summaries:
    body = msg['body']
    msg_date = datetime.strptime(msg['date'], "%d/%m/%Y")

    # Try to extract the actual report date from body
    report_date = None
    # "25th February 2026" or "25 February 2026"
    rd_match = re.search(r'(\d{1,2})(?:st|nd|rd|th)?\s+(January|February|March|April|May|June|July|August|September|October|November|December)\s+(\d{4})', body)
    if rd_match:
        try:
            report_date = datetime.strptime(f"{rd_match.group(1)} {rd_match.group(2)} {rd_match.group(3)}", "%d %B %Y")
        except:
            pass
    if not report_date:
        # DD.MM.YYYY
        rd_match2 = re.search(r'(\d{2}\.\d{2}\.\d{4})', body)
        if rd_match2:
            try:
                report_date = datetime.strptime(rd_match2.group(1), "%d.%m.%Y")
            except:
                pass
    if not report_date:
        report_date = msg_date

    # Shift
    shift = "Unknown"
    if re.search(r'[Nn]ight\s*[Ss]hift', body):
        shift = "Night"
    elif re.search(r'[Dd]ay\s*[Ss]hift', body):
        shift = "Day"

    # Team
    team_match = re.search(r'[Tt]eam\s*[:\-]?\s*([A-D])', body)
    team = team_match.group(1) if team_match else ""

    # Machine data: look for each machine ID followed by hours and/or loads
    # We need to parse blocks like:
    # *KMT2503:* 2353.7 - 2363.9
    # Fuel: 83%
    # Loads: 50
    machine_ids = {
        'KME2203': 'FEL',
        'KME2215': 'FEL',
        'KME2510': 'EXC',
        'KME2509': 'Roller',
        'KME2603': 'Dozer',
        'KMT2503': 'ADT',
        'KMT2504': 'ADT',
        'KMT2505': 'ADT',
        'KMT2506': 'ADT',
        'KMT2507': 'ADT',
        'C030': 'Roller',
        'R21SA': 'Roller',
        'C095': 'ADT (Lalgy)',
        'C096': 'ADT (Lalgy)',
        'C105': 'ADT (Lalgy)',
    }

    machines = {}
    for mid, mtype in machine_ids.items():
        # Check if machine is mentioned
        if mid not in body:
            continue

        # Check status - is it down/breakdown/no operator?
        # Look for the machine ID and the text near it
        machine_block_match = re.search(
            rf'{mid}[^A-Z]*?((?:On\s*[Bb]reakdown|[Nn]o\s*[Oo]perator|[Ss]tanding|[Dd]id\s*not\s*run|[Oo]n\s*standby|[Ss]ervice|[Ff]ire\s*extinguisher))',
            body, re.IGNORECASE
        )
        status = 'down' if machine_block_match else 'ok'

        # Extract loads (only for ADTs)
        loads = None
        if 'ADT' in mtype or 'Lalgy' in mtype:
            # Find the machine ID and then look for "Loads:" nearby
            # Strategy: find all occurrences of the machine id, then look for loads after each
            pattern = re.compile(
                rf'{mid}.*?[Ll]oads[:\s=]*(\d+)',
                re.DOTALL
            )
            # But limit the search to avoid crossing into the next machine
            # Find position of this machine
            pos = body.find(mid)
            if pos >= 0:
                # Find next machine or end
                next_pos = len(body)
                for other_mid in machine_ids:
                    if other_mid == mid:
                        continue
                    p = body.find(other_mid, pos + len(mid))
                    if p >= 0 and p < next_pos:
                        next_pos = p
                # Also look for "Total" as a boundary
                total_pos = body.lower().find('total', pos + len(mid))
                if total_pos >= 0 and total_pos < next_pos:
                    next_pos = total_pos

                block = body[pos:next_pos]
                load_match = re.search(r'[Ll]oads[:\s=]*(\d+)', block)
                if load_match:
                    loads = int(load_match.group(1))

        # Extract hours
        hours_start = None
        hours_end = None
        pos = body.find(mid)
        if pos >= 0:
            next_pos = len(body)
            for other_mid in machine_ids:
                if other_mid == mid:
                    continue
                p = body.find(other_mid, pos + len(mid))
                if p >= 0 and p < next_pos:
                    next_pos = p
            block = body[pos:next_pos]
            # Look for hour patterns: "1234.5 - 1245.6" or just "1234.5"
            hours_match = re.search(r'(\d{3,5}[\.,]\d+)\s*[-–]\s*(\d{3,5}[\.,]\d+)', block)
            if hours_match:
                hours_start = hours_match.group(1).replace(',', '.')
                hours_end = hours_match.group(2).replace(',', '.')
            else:
                hours_match2 = re.search(r'[:\s](\d{3,5}[\.,]\d+)', block)
                if hours_match2:
                    hours_end = hours_match2.group(1).replace(',', '.')

        machines[mid] = {
            'type': mtype,
            'status': status,
            'loads': loads,
            'hours_start': hours_start,
            'hours_end': hours_end,
        }

    # Total loads
    total_match = re.search(r'[Tt]otal\s*(?:de\s*)?[Ll]oads[:\s=*\-]*(\d+)', body)
    total_loads = int(total_match.group(1)) if total_match else None

    rows.append({
        'date': report_date.strftime("%Y-%m-%d"),
        'shift': shift,
        'team': team,
        'machines': machines,
        'total_loads': total_loads,
        'sender': msg['sender'],
    })

# Filter out bad dates (before 2025 or after 2027)
rows = [r for r in rows if r['date'] >= '2025-01-01' and r['date'] <= '2027-01-01']

# Sort
rows.sort(key=lambda r: (r['date'], 0 if r['shift'] == 'Day' else 1 if r['shift'] == 'Night' else 2))

# ---- Build Excel workbook ----
wb = Workbook()
ws = wb.active
ws.title = "Lift 1 Production"

# Styles
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=11)
kenmare_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
lalgy_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
down_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
down_font = Font(color="990000")
thin_border = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

# Kenmare machines (ADTs)
kenmare_adts = ['KMT2503', 'KMT2504', 'KMT2505', 'KMT2506', 'KMT2507']
# Lalgy machines
lalgy_adts = ['C095', 'C096', 'C105']
# Support machines
support = ['KME2203', 'KME2510', 'KME2509', 'KME2603', 'C030', 'R21SA']

# Headers
headers = ['Date', 'Shift', 'Team']
# Kenmare ADT loads
for mid in kenmare_adts:
    headers.append(f'{mid}')
headers.append('Kenmare Total')
# Lalgy ADT loads
for mid in lalgy_adts:
    headers.append(f'{mid}')
headers.append('Lalgy Total')
headers.append('Grand Total')
# Support machine status
for mid in support:
    headers.append(mid)

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Subheader row for machine types
ws.insert_rows(2)
type_labels = {
    'KMT2503': 'ADT', 'KMT2504': 'ADT', 'KMT2505': 'ADT', 'KMT2506': 'ADT', 'KMT2507': 'ADT',
    'C095': 'ADT', 'C096': 'ADT', 'C105': 'ADT',
    'KME2203': 'FEL', 'KME2510': 'EXC', 'KME2509': 'Roller', 'KME2603': 'Dozer',
    'C030': 'Roller', 'R21SA': 'Roller',
}
sub_labels = ['', '', '']
for mid in kenmare_adts:
    sub_labels.append('Loads')
sub_labels.append('')
for mid in lalgy_adts:
    sub_labels.append('Loads')
sub_labels.append('')
sub_labels.append('')
for mid in support:
    sub_labels.append(type_labels.get(mid, ''))

for col_idx, label in enumerate(sub_labels, 1):
    cell = ws.cell(row=2, column=col_idx, value=label)
    cell.font = Font(italic=True, size=9, color="666666")
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

# Data rows
for row_idx, row_data in enumerate(rows, 3):
    machines = row_data['machines']

    ws.cell(row=row_idx, column=1, value=row_data['date']).border = thin_border
    ws.cell(row=row_idx, column=2, value=row_data['shift']).border = thin_border
    ws.cell(row=row_idx, column=3, value=row_data['team']).border = thin_border

    col = 4
    kenmare_total = 0
    kenmare_count = 0

    # Kenmare ADTs
    for mid in kenmare_adts:
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        if mid in machines:
            m = machines[mid]
            if m['status'] == 'down':
                cell.value = m.get('loads', 'DOWN')
                if cell.value == 'DOWN' or cell.value is None:
                    cell.value = 'DOWN'
                cell.fill = down_fill
                cell.font = down_font
            else:
                loads = m.get('loads')
                if loads is not None:
                    cell.value = loads
                    kenmare_total += loads
                    kenmare_count += 1
                else:
                    cell.value = '—'
        else:
            cell.value = ''
        col += 1

    # Kenmare total
    cell = ws.cell(row=row_idx, column=col, value=kenmare_total if kenmare_count > 0 else '')
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = kenmare_fill
    col += 1

    # Lalgy ADTs
    lalgy_total = 0
    lalgy_count = 0
    for mid in lalgy_adts:
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        if mid in machines:
            m = machines[mid]
            if m['status'] == 'down':
                cell.value = 'DOWN'
                cell.fill = down_fill
                cell.font = down_font
            else:
                loads = m.get('loads')
                if loads is not None:
                    cell.value = loads
                    lalgy_total += loads
                    lalgy_count += 1
                else:
                    cell.value = '—'
        else:
            cell.value = ''
        col += 1

    # Lalgy total
    cell = ws.cell(row=row_idx, column=col, value=lalgy_total if lalgy_count > 0 else '')
    cell.border = thin_border
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')
    cell.fill = lalgy_fill
    col += 1

    # Grand total
    total = row_data.get('total_loads')
    cell = ws.cell(row=row_idx, column=col, value=total if total else '')
    cell.border = thin_border
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal='center')
    col += 1

    # Support machines status
    for mid in support:
        cell = ws.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

        if mid in machines:
            m = machines[mid]
            if m['status'] == 'down':
                cell.value = 'DOWN'
                cell.fill = down_fill
                cell.font = down_font
            else:
                # Show closing hours if available
                if m.get('hours_end'):
                    cell.value = float(m['hours_end'])
                else:
                    cell.value = 'OK'
        else:
            cell.value = ''
        col += 1

# Column widths
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 6
for col_idx in range(4, len(headers) + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 11

# Freeze panes
ws.freeze_panes = 'D3'

# Save
outfile = "../wiki/data/lift1-production.xlsx"
wb.save(outfile)
print(f"Exported {len(rows)} shift summaries to {outfile}")
